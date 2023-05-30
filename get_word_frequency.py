import re
import requests
import openpyxl
import hashlib
import random
from operator import itemgetter


IGNORE_WORDS_FILE = 'ignore_words.txt'
TEXT_FILE = 'text.txt'
OUTPUT_FILE_TEXT = 'output.txt'
OUTPUT_FILE_EXCEL = 'output.xlsx'
YOUDAO_API_URL = 'http://fanyi.youdao.com/translate?doctype=json&type=EN2ZH_CN&i='
BAIDU_API_URL = 'http://api.fanyi.baidu.com/api/trans/vip/translate?q={word}&from=en&to=zh&appid={appid}&salt={salt}&sign={sign}'
BAIDU_APPID = 'YOUR_BAIDU_APPID'
BAIDU_SECRET_KEY = 'YOUR_BAIDU_SECRET_KEY'


def preprocess_text(text):
    # 去除标点符号和转换为小写
    text = re.sub(r'[^\w\s]', '', text.lower())
    # 拆分文本为单词列表
    words = text.split()
    return words


def get_word_frequency(words, ignore_words):
    # 创建一个字典用于存储单词及其频率
    word_frequency = {}

    for word in words:
        if word not in ignore_words:
            if word in word_frequency:
                word_frequency[word] += 1
            else:
                word_frequency[word] = 1

    return word_frequency


def get_word_translation_youdao(word):
    url = YOUDAO_API_URL + word
    response = requests.get(url)
    data = response.json()

    if response.status_code == 200:
        # 解析响应并获取中文解释
        translations = data['translateResult'][0][0]['tgt']
        return translations
    else:
        return None


def get_word_translation_baidu(word):
    salt = str(random.randint(32768, 65536))
    sign = hashlib.md5((BAIDU_APPID + word + salt +
                       BAIDU_SECRET_KEY).encode()).hexdigest()
    url = BAIDU_API_URL.format(
        word=word, appid=BAIDU_APPID, salt=salt, sign=sign)

    response = requests.get(url)
    data = response.json()

    if 'trans_result' in data:
        translation = data['trans_result'][0]['dst']
        return translation
    else:
        return None


def save_to_text_file(sorted_word_frequency, dict_option):
    with open(OUTPUT_FILE_TEXT, 'w', encoding='utf-8') as file:
        # 写入标题行
        file.write("单词, 频率, 中文解释\n")

        for word, frequency in sorted_word_frequency:
            file.write(f"{word}, {frequency}, ")

            # 获取中文解释
            if dict_option == "1":
                translation = get_word_translation_youdao(word)
            elif dict_option == "2":
                translation = get_word_translation_baidu(word)
            else:
                translation = get_word_translation_youdao(word)

            if translation:
                file.write(f"{translation}\n")
            else:
                file.write("无法获取中文解释\n")


def save_to_excel_file(sorted_word_frequency, dict_option):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # 写入标题行
    sheet.cell(row=1, column=1, value="单词")
    sheet.cell(row=1, column=2, value="频率")
    sheet.cell(row=1, column=3, value="中文解释")

    row_num = 2
    for word, frequency in sorted_word_frequency:
        sheet.cell(row=row_num, column=1, value=word)
        sheet.cell(row=row_num, column=2, value=frequency)

        # 获取中文解释
        if dict_option == "1":
            translation = get_word_translation_youdao(word)
        elif dict_option == "2":
            translation = get_word_translation_baidu(word)
        else:
            translation = get_word_translation_youdao(word)

        if translation:
            sheet.cell(row=row_num, column=3, value=translation)
        else:
            sheet.cell(row=row_num, column=3, value="无法获取中文解释")

        row_num += 1

    workbook.save(OUTPUT_FILE_EXCEL)


def main():
    # 从文本文件中读取英文文本
    with open(TEXT_FILE, 'r', encoding='utf-8') as file:
        text = file.read()

    # 从文件中读取需要忽略的词汇
    with open(IGNORE_WORDS_FILE, 'r', encoding='utf-8') as file:
        ignore_words = [line.strip() for line in file]

    # 预处理文本
    words = preprocess_text(text)

    # 获取单词频率
    word_frequency = get_word_frequency(words, ignore_words)

    # 按单词频率降序排序
    sorted_word_frequency = sorted(
        word_frequency.items(), key=itemgetter(1), reverse=True)

    # 选择输出选项
    output_option = input("请选择输出选项（输入1代表文本文件，输入2代表Excel文件）：")
    dict_option = input("请选择使用哪个词典（输入1代表有道词典，输入2代表百度词典）：")

    if output_option == "1":
        save_to_text_file(sorted_word_frequency, dict_option)
    else:
        save_to_excel_file(sorted_word_frequency, dict_option)


if __name__ == '__main__':
    main()
